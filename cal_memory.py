import openpyxl

files = ['Hirain', 'CTASEN', 'DMC', 'EM', 'LBSSEN', 'LCFSEN', 'LCFVEH', 'VED', 'VLCSEN', 'VLCVEH', 'ALL']
mp_file = {'Hirain' : 'EP40_ADCU_hirain_src.map', 
           'CTASEN' : 'EP40_ADCU_CTASEN.map',
           'DMC' : 'EP40_ADCU_DMC.map',
           'EM' : 'EP40_ADCU_EM.map',
           'LBSSEN' : 'EP40_ADCU_LBSSEN.map',
           'LCFSEN' : 'EP40_ADCU_LCFSEN.map',
           'LCFVEH' : 'EP40_ADCU_LCFVEH.map',
           'VED' : 'EP40_ADCU_VED.map',
           'VLCSEN' : 'EP40_ADCU_VLCSEN.map',
           'VLCVEH' : 'EP40_ADCU_VLCVEH.map',
           'ALL' : 'EP40_ADCU_ALL.map'}

mp = {} # mp[file][i]是一个列表，表示file第i行的内存名、free和total

for i in range(len(files)):
    filename = mp_file[files[i]]
    mp[files[i]] = []
    with open(filename, 'r') as f:
        lines = f.readlines()
    index = []
    for j, line in enumerate(lines):
        if line == '+-----------------------------------------------------------------------------------------------------+\n':
            index.append(j)
            
    for line in lines[index[0] + 3 : index[1]]:
        if line == '|-----------------------------------------------------------------------------------------------------|\n':
            continue
        temp = []
        element = line.split()
        # print(element)
        item = element[1]
        free = int(element[9], 16)
        total = int(element[11], 16)
        mp[files[i]].append([item, free, total])
        
        
book = openpyxl.Workbook()

# 新建sheet
book.create_sheet(title=r'原始数据')
book.create_sheet(title=r'内存使用数据')
book.create_sheet(title=r'内存使用数据(Kb)')
book.remove(book['Sheet'])  #删除默认sheet

# 第一张表
sheet1 = book[r'原始数据']
first_line = ['MEM', 'Hirain', 'CTASEN', 'DMC', 'EM', 'LBSSEN', 'LCFSEN', 'LCFVEH', 'VED', 'VLCSEN', 'VLCVEH', 'ALL', 'TOTAL']
second_line = ['Memory', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Total']

# 写前两行
for i in range(len(first_line)):
    sheet1.cell(1, column=i+1).value = first_line[i]
    sheet1.cell(2, column=i+1).value = second_line[i]
    
# 写第一列和最后一列
for i in range(51):
    row = i+3
    sheet1.cell(row, 1).value = mp[files[0]][i][0]
    sheet1.cell(row, 13).value = mp[files[0]][i][2]
    
# 写其他
for i in range(51):
    for j in range(11):
        row = i+3
        column = j+2
        sheet1.cell(row, column).value = mp[files[j]][i][1]

# 第二、三张表
sheet2 = book[r'内存使用数据']

first_line = ['MEM', 'Hirain', 'CTASEN', 'DMC', 'EM', 'LBSSEN', 'LCFSEN', 'LCFVEH', 'VED', 'VLCSEN', 'VLCVEH', r'所有使用内存', r'剩余内', 'TOTAL']
second_line = ['Memory', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Free', 'Total']

#写第一行
for i in range(len(first_line)):
    sheet2.cell(1, column=i+1).value = first_line[i]
    
sheet2.cell(2, 1).value = 'Total RAM'

# 写第1,13,14列
for i in range(3, 54):
    sheet2.cell(i, 1).value = sheet1.cell(i, 1).value
    sheet2.cell(i, 13).value = sheet1.cell(i, 12).value
    sheet2.cell(i, 14).value = sheet1.cell(i, 13).value    
    
#写第2列
for i in range(3, 54):
    sheet2.cell(i, 2).value = sheet1.cell(i, 13).value - sheet1.cell(i, 2).value
    
#写其他数
for i in range(3, 54):
    sheet2.cell(i, 12).value = sheet1.cell(i, 13).value - sheet1.cell(i, 12).value
    for j in range(3, 12):
        sheet2.cell(i, j).value = sheet1.cell(i, 2).value - sheet1.cell(i, j).value

#写第2行
for j in range(2, 15):
    sheet2.cell(2, j).value = sheet2.cell(53, j).value - sheet2.cell(51, j).value

# 第三张表，把第二张表复制一下，数值除以1024
sheet3 = book[r'内存使用数据(Kb)']

for i in range(1, 54):
    for j in range(1, 15):
        sheet3.cell(i, j).value = sheet2.cell(i, j).value
        if not isinstance(sheet3.cell(i, j).value, str):
            sheet3.cell(i, j).value /= 1024
   
# 保存表
book.save('memory_statistics.xlsx')
